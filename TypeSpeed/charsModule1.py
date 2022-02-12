import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from ctypes import windll

def charsMod():	
    
    name=input("Enter Your Name : ")

    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()

    if len(name)==0:
        print("\nINVALID NAME!")
        exit()
    else:
        for n in namelist:
            if n==name:
                print("\nName Already Exists\n")
                exit()
    
    n=int(input("\nTo Start Typing, Enter 1 and <ENTER> and Start Typing..\n\n"))
    
    if n==1:
    
        if windll.user32.OpenClipboard(None):
            windll.user32.EmptyClipboard()
            windll.user32.CloseClipboard()
        
        s=time.time()
        str=input("")
        string=str.lower()
    
        list=[c for c in string]    
        n=len(list)
        
        if n==0:
            print("\nINVALID INPUT")
        elif string.isalpha()==False:
            print("\nINVALID CHARACHTER!")
            exit()
        elif n>26:
            print("\nEXTRA CHARACHTERS!")
            exit()
    
        for i in range(len(list)):
            if list[0]=="a":
                if n<2:
                    print("\nNo Charachters after A!")
                    e=time.time()
                    print("Time is : ",e-s)
                    quit()
                else:
                    if list[1]=="b":
                        if n<3:
                            print("\nNo Charachters after B!")
                            e=time.time()
                            print("Time is : ",e-s)
                            quit()
                        else:
                            if list[2]=="c":
                                if n<4:
                                    print("\nNo Charachters after C!")
                                    e=time.time()
                                    print("Time is : ",e-s)
                                    quit()
                                else:
                                    if list[3]=="d":
                                        if n<5:
                                            print("\nNo Charachters after D!")
                                            e=time.time()
                                            print("Time is : ",e-s)
                                            quit()
                                        else:
                                            if list[4]=="e":
                                                if n<6:
                                                    print("\nNo Charachters after E!")
                                                    e=time.time()
                                                    print("Time is : ",e-s)
                                                    quit()
                                                else:
                                                    if list[5]=="f":
                                                        if n<7:
                                                            print("\nNo Charachters after F!")
                                                            e=time.time()
                                                            print("Time is : ",e-s)
                                                            quit()
                                                        else:
                                                            if list[6]=="g":
                                                                if n<8:
                                                                    print("\nNo Charachters after G!")
                                                                    e=time.time()
                                                                    print("Time is : ",e-s)
                                                                    quit()
                                                                else:
                                                                    if list[7]=="h":
                                                                        if n<9:
                                                                            print("\nNo Charachters after H!")
                                                                            e=time.time()
                                                                            print("Time is : ",e-s)
                                                                            quit()
                                                                        else:
                                                                            if list[8]=="i":
                                                                                if n<10:
                                                                                    print("\nNo Charachters after I!")
                                                                                    e=time.time()
                                                                                    print("Time is : ",e-s)
                                                                                    quit()
                                                                                else:
                                                                                    if list[9]=="j":
                                                                                        if n<11:
                                                                                            print("\nNo Charachters after J!")
                                                                                            e=time.time()
                                                                                            print("Time is : ",e-s)
                                                                                            quit()
                                                                                        else:
                                                                                            if list[10]=="k":
                                                                                                if n<12:
                                                                                                    print("\nNo Charachters after K!")
                                                                                                    e=time.time()
                                                                                                    print("Time is : ",e-s)
                                                                                                    quit()
                                                                                                else:
                                                                                                    if list[11]=="l":
                                                                                                        if n<13:
                                                                                                            print("\nNo Charachters after L!")
                                                                                                            e=time.time()
                                                                                                            print("Time is : ",e-s)
                                                                                                            quit()
                                                                                                        else:
                                                                                                            if list[12]=="m":
                                                                                                                if n<14:
                                                                                                                    print("\nNo Charachters after M!")
                                                                                                                    e=time.time()
                                                                                                                    print("Time is : ",e-s)
                                                                                                                    quit()
                                                                                                                else:
                                                                                                                    if list[13]=="n":
                                                                                                                        if n<15:
                                                                                                                            print("\nNo Charachters after N!")
                                                                                                                            e=time.time()
                                                                                                                            print("Time is : ",e-s)
                                                                                                                            quit()
                                                                                                                        else:
                                                                                                                            if list[14]=="o":
                                                                                                                                if n<16:
                                                                                                                                    print("\nNo Charachters after O!")
                                                                                                                                    e=time.time()
                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                    quit()
                                                                                                                                else:
                                                                                                                                    if list[15]=="p":
                                                                                                                                        if n<17:
                                                                                                                                            print("\nNo Charachters after P!")
                                                                                                                                            e=time.time()
                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                            quit()
                                                                                                                                        else:
                                                                                                                                            if list[16]=="q":
                                                                                                                                                if n<18:
                                                                                                                                                    print("\nNo Charachters after Q!")
                                                                                                                                                    e=time.time()
                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                    quit()
                                                                                                                                                else:
                                                                                                                                                    if list[17]=="r":
                                                                                                                                                        if n<19:
                                                                                                                                                            print("\nNo Charachters after R!")
                                                                                                                                                            e=time.time()
                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                            quit()
                                                                                                                                                        else:
                                                                                                                                                            if list[18]=="s":
                                                                                                                                                                if n<20:
                                                                                                                                                                    print("\nNo Charachters after S!")
                                                                                                                                                                    e=time.time()
                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                    quit()
                                                                                                                                                                else:
                                                                                                                                                                    if list[19]=="t":
                                                                                                                                                                        if n<21:
                                                                                                                                                                            print("\nNo Charachters after T!")
                                                                                                                                                                            e=time.time()
                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                            quit()
                                                                                                                                                                        else:
                                                                                                                                                                            if list[20]=="u":
                                                                                                                                                                                if n<22:
                                                                                                                                                                                    print("\nNo Charachters after U!")
                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                    quit()
                                                                                                                                                                                else:
                                                                                                                                                                                    if list[21]=="v":
                                                                                                                                                                                        if n<23:
                                                                                                                                                                                            print("\nNo Charachters after V!")
                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                            quit()
                                                                                                                                                                                        else:
                                                                                                                                                                                            if list[22]=="w":
                                                                                                                                                                                                if n<24:
                                                                                                                                                                                                    print("\nNo Charachters after W!")
                                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                                    quit()
                                                                                                                                                                                                else:
                                                                                                                                                                                                    if list[23]=="x":
                                                                                                                                                                                                        if n<25:
                                                                                                                                                                                                            print("\nNo Charachters after X!")
                                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                                            quit()
                                                                                                                                                                                                        else:
                                                                                                                                                                                                            if list[24]=="y":
                                                                                                                                                                                                                if n<26:
                                                                                                                                                                                                                    print("\nNo Charachter after Y!")
                                                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                                                    quit()
                                                                                                                                                                                                                else:
                                                                                                                                                                                                                    if list[25]=="z":
                                                                                                                                                                                                                        print("\nSuccess..")
                                                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                                                        t=e-s
                                                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                                                        print("WPM : ",(26/5)/((e-s)/60))
                                                                                                                                                                                                                        print()
                                                                                                                                                                                                                        ld(name,t)
                                                                                                                                                                                                                        exit()
                                                                                                                                                                                                                    else:
                                                                                                                                                                                                                        print("\nZ Misplaced!")
                                                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                                                        quit()
                                                                                                                                                                                                            else:
                                                                                                                                                                                                                print("\nY Misplaced!")
                                                                                                                                                                                                                e=time.time()
                                                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                                                quit()
                                                                                                                                                                                                    else:
                                                                                                                                                                                                        print("\nX Misplaced!")
                                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                                        quit()
                                                                                                                                                                                            else:
                                                                                                                                                                                                print("\nW Misplaced!")
                                                                                                                                                                                                e=time.time()
                                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                                quit()
                                                                                                                                                                                    else:
                                                                                                                                                                                        print("\nV Misplaced!")
                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                        quit()
                                                                                                                                                                            else:
                                                                                                                                                                                print("\nU Misplaced!")
                                                                                                                                                                                e=time.time()
                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                quit()
                                                                                                                                                                    else:
                                                                                                                                                                        print("\nT Misplaced!")
                                                                                                                                                                        e=time.time()
                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                        quit()
                                                                                                                                                            else:
                                                                                                                                                                print("\nS Misplaced!")
                                                                                                                                                                e=time.time()
                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                quit()
                                                                                                                                                    else:
                                                                                                                                                        print("\nR Misplaced!")
                                                                                                                                                        e=time.time()
                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                        quit()
                                                                                                                                            else:
                                                                                                                                                print("\nQ Misplaced!")
                                                                                                                                                e=time.time()
                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                quit()
                                                                                                                                    else:
                                                                                                                                        print("\nP Misplaced!")
                                                                                                                                        e=time.time()
                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                        quit()
                                                                                                                            else:
                                                                                                                                print("\nO Misplaced!")
                                                                                                                                e=time.time()
                                                                                                                                print("Time is : ",e-s)
                                                                                                                                quit()
                                                                                                                    else:
                                                                                                                        print("\nN Misplaced!")
                                                                                                                        e=time.time()
                                                                                                                        print("Time is : ",e-s)
                                                                                                                        quit()
                                                                                                            else:
                                                                                                                print("\nM Misplaced!")
                                                                                                                e=time.time()
                                                                                                                print("Time is : ",e-s)
                                                                                                                quit()
                                                                                                    else:
                                                                                                        print("\nL Misplaced!")
                                                                                                        e=time.time()
                                                                                                        print("Time is : ",e-s)
                                                                                                        quit()
                                                                                            else:
                                                                                                print("\nK Misplaced!")
                                                                                                e=time.time()
                                                                                                print("Time is : ",e-s)
                                                                                                quit()
                                                                                    else:
                                                                                        print("\nJ Misplaced!")
                                                                                        e=time.time()
                                                                                        print("Time is : ",e-s)
                                                                                        quit()
                                                                            else:
                                                                                print("\nI Misplaced!")
                                                                                e=time.time()
                                                                                print("Time is : ",e-s)
                                                                                quit()
                                                                    else:
                                                                        print("\nH Misplaced!")
                                                                        e=time.time()
                                                                        print("Time is : ",e-s)
                                                                        quit()
                                                            else:
                                                                print("\nG Misplaced!")
                                                                e=time.time()
                                                                print("Time is : ",e-s)
                                                                quit()
                                                    else:
                                                        print("\nF Misplaced!")
                                                        e=time.time()
                                                        print("Time is : ",e-s)
                                                        quit()
                                            else:
                                                print("\nE Misplaced!")
                                                e=time.time()
                                                print("Time is : ",e-s)
                                                quit()
                                    else:
                                        print("\nD Misplaced!")
                                        e=time.time()
                                        print("Time is : ",e-s)
                                        quit()
                            else:
                                print("\nC Misplaced!")
                                e=time.time()
                                print("Time is : ",e-s)
                                quit()
                    else:
                        print("\nB Misplaced!")
                        e=time.time()
                        print("Time is : ",e-s)
                        quit()
            else:
                print("\nA Misplaced!")
                e=time.time()
                print("Time is : ",e-s)
                quit()
    else:
        
        print("\nINVALID INPUT")
        
def ld(n1,t1):

    name=n1
    time=t1
    dict={}
    sdict={}
    namelist=[]
    timelist=[]
    
    ob=openpyxl.load_workbook("TypeBoard.xlsx")
    ba=ob.active
    
    for i in range(1,100):
        if ba.cell(row=i,column=1).value is None and ba.cell(row=i,column=2).value is None:
            ba.cell(row=i,column=1).value=name
            ba.cell(row=i,column=2).value=time
            break
    
    ob.save(filename="TypeBoard.xlsx")
    
    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()
    timelist=data["Time"].tolist()

    for i in range(len(namelist)):
        dict[namelist[i]]=timelist[i]
    
    sortv=sorted(dict.values())
    
    for i in sortv:
        for j in dict.keys():
            if dict[j]==i:
                sdict[j]=dict[j]
                break
                
    print("\nTOP 5\n")
    
    i=0
    for key in sdict:
        if i!=5:
            print(key,":",sdict[key])
            i=i+1    
