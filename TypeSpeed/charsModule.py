import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
import pyperclip

def charsMod():	
    
    alph="abcdefghijklmnopqrstuvwxyz"
    
    name=input("Enter Your Name : ")

    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()
    
    if len(name)==0:
        print("\nINVALID NAME!")
        exit()
    else:
        for n in namelist:
            if n==name:
                print("\nName Already Exists\n")
                exit()
    
    n=int(input("\nTo Start Typing, Enter 1 and <ENTER> and Start Typing..\n\n"))
    
    if n==1:
    
        pyperclip.copy('')
        s=time.time()
        str=input("")
        string=str.lower()
    
        n=len(string)
        
        if n==0:
            print("\nINVALID INPUT")
        elif string.isalpha()==False:
            print("\nINVALID CHARACHTER!")
            exit()
        elif n>26:
            print("\nEXTRA CHARACHTERS!")
            exit()
    
        if string==alph:
            print("\nSuccess..")
            e=time.time()
            t=e-s
            print("Time is : ",e-s)
            print("WPM : ",(26/5)/((e-s)/60))
            print()
            ld(name,t)
            exit()
        else:
            print("\nWRONG CHARACHTER")
            e=time.time()
            t=e-s
            print("Time is : ",e-s)

    else:
        
        print("\nINVALID INPUT")
        
def ld(n1,t1):

    name=n1
    time=t1
    dict={}
    sdict={}
    namelist=[]
    timelist=[]
    
    ob=openpyxl.load_workbook("TypeBoard.xlsx")
    ba=ob.active
    
    for i in range(1,100):
        if ba.cell(row=i,column=1).value is None and ba.cell(row=i,column=2).value is None:
            ba.cell(row=i,column=1).value=name
            ba.cell(row=i,column=2).value=time
            break
    
    ob.save(filename="TypeBoard.xlsx")
    
    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()
    timelist=data["Time"].tolist()

    for i in range(len(namelist)):
        dict[namelist[i]]=timelist[i]
    
    sortv=sorted(dict.values())
    
    for i in sortv:
        for j in dict.keys():
            if dict[j]==i:
                sdict[j]=dict[j]
                break
                
    print("\nTOP 5\n")
    
    i=0
    for key in sdict:
        if i!=5:
            print(key,":",sdict[key])
            i=i+1    
