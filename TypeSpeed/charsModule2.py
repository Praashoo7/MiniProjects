import time
import pandas as pd
import openpyxl
from openpyxl import Workbook

def charsMod():	
    
    name=input("Enter Your Name : ")

    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()
    
    for n in namelist:
        if n==name:
            print("\nName Already Exists\n")
            exit()
    
    n=int(input("\nTo Start Typing, Enter 1 and <ENTER> and Start Typing..\n\n"))
    
    if n==1:
    
        s=time.time()
        string=input("")
    
        list=[c for c in string]    
        n=len(list)
        
        if n==0:
            print("\nINVALID INPUT")
    
        for i in range(len(list)):
            if list[0]=="A" or list[0]=="a":
                if n<2:
                    print("\nWRONG CAHARCHTER")
                    e=time.time()
                    print("Time is : ",e-s)
                    quit()
                else:
                    if list[1]=="B" or list[1]=="b":
                        if n<3:
                            print("\nWRONG CAHARCHTER")
                            e=time.time()
                            print("Time is : ",e-s)
                            quit()
                        else:
                            if list[2]=="C" or list[2]=="c":
                                if n<4:
                                    print("\nWRONG CAHARCHTER")
                                    e=time.time()
                                    print("Time is : ",e-s)
                                    quit()
                                else:
                                    if list[3]=="D" or list[3]=="d":
                                        if n<5:
                                            print("\nWRONG CAHARCHTER")
                                            e=time.time()
                                            print("Time is : ",e-s)
                                            quit()
                                        else:
                                            if list[4]=="E" or list[4]=="e":
                                                if n<6:
                                                    print("\nWRONG CAHARCHTER")
                                                    e=time.time()
                                                    print("Time is : ",e-s)
                                                    quit()
                                                else:
                                                    if list[5]=="F" or list[5]=="f":
                                                        if n<7:
                                                            print("\nWRONG CAHARCHTER")
                                                            e=time.time()
                                                            print("Time is : ",e-s)
                                                            quit()
                                                        else:
                                                            if list[6]=="G" or list[6]=="g":
                                                                if n<8:
                                                                    print("\nWRONG CAHARCHTER")
                                                                    e=time.time()
                                                                    print("Time is : ",e-s)
                                                                    quit()
                                                                else:
                                                                    if list[7]=="H" or list[7]=="h":
                                                                        if n<9:
                                                                            print("\nWRONG CAHARCHTER")
                                                                            e=time.time()
                                                                            print("Time is : ",e-s)
                                                                            quit()
                                                                        else:
                                                                            if list[8]=="I" or list[8]=="i":
                                                                                if n<10:
                                                                                    print("\nWRONG CAHARCHTER")
                                                                                    e=time.time()
                                                                                    print("Time is : ",e-s)
                                                                                    quit()
                                                                                else:
                                                                                    if list[9]=="J" or list[9]=="j":
                                                                                        if n<11:
                                                                                            print("\nWRONG CAHARCHTER")
                                                                                            e=time.time()
                                                                                            print("Time is : ",e-s)
                                                                                            quit()
                                                                                        else:
                                                                                            if list[10]=="K" or list[10]=="k":
                                                                                                if n<12:
                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                    e=time.time()
                                                                                                    print("Time is : ",e-s)
                                                                                                    quit()
                                                                                                else:
                                                                                                    if list[11]=="L" or list[11]=="l":
                                                                                                        if n<13:
                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                            e=time.time()
                                                                                                            print("Time is : ",e-s)
                                                                                                            quit()
                                                                                                        else:
                                                                                                            if list[12]=="M" or list[12]=="m":
                                                                                                                if n<14:
                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                    e=time.time()
                                                                                                                    print("Time is : ",e-s)
                                                                                                                    quit()
                                                                                                                else:
                                                                                                                    if list[13]=="N" or list[13]=="n":
                                                                                                                        if n<15:
                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                            e=time.time()
                                                                                                                            print("Time is : ",e-s)
                                                                                                                            quit()
                                                                                                                        else:
                                                                                                                            if list[14]=="O" or list[14]=="o":
                                                                                                                                if n<16:
                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                    e=time.time()
                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                    quit()
                                                                                                                                else:
                                                                                                                                    if list[15]=="P" or list[15]=="p":
                                                                                                                                        if n<17:
                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                            e=time.time()
                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                            quit()
                                                                                                                                        else:
                                                                                                                                            if list[16]=="Q" or list[16]=="q":
                                                                                                                                                if n<18:
                                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                                    e=time.time()
                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                    quit()
                                                                                                                                                else:
                                                                                                                                                    if list[17]=="R" or list[17]=="r":
                                                                                                                                                        if n<19:
                                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                                            e=time.time()
                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                            quit()
                                                                                                                                                        else:
                                                                                                                                                            if list[18]=="S" or list[18]=="s":
                                                                                                                                                                if n<20:
                                                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                                                    e=time.time()
                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                    quit()
                                                                                                                                                                else:
                                                                                                                                                                    if list[19]=="T" or list[19]=="t":
                                                                                                                                                                        if n<21:
                                                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                                                            e=time.time()
                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                            quit()
                                                                                                                                                                        else:
                                                                                                                                                                            if list[20]=="U" or list[20]=="u":
                                                                                                                                                                                if n<22:
                                                                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                    quit()
                                                                                                                                                                                else:
                                                                                                                                                                                    if list[21]=="V" or list[21]=="v":
                                                                                                                                                                                        if n<23:
                                                                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                            quit()
                                                                                                                                                                                        else:
                                                                                                                                                                                            if list[22]=="W" or list[22]=="w":
                                                                                                                                                                                                if n<24:
                                                                                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                                    quit()
                                                                                                                                                                                                else:
                                                                                                                                                                                                    if list[23]=="X" or list[23]=="x":
                                                                                                                                                                                                        if n<25:
                                                                                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                                            quit()
                                                                                                                                                                                                        else:
                                                                                                                                                                                                            if list[24]=="Y" or list[24]=="y":
                                                                                                                                                                                                                if n<26:
                                                                                                                                                                                                                    print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                                    e=time.time()
                                                                                                                                                                                                                    print("Time is : ",e-s)
                                                                                                                                                                                                                    quit()
                                                                                                                                                                                                                else:
                                                                                                                                                                                                                    if list[25]=="Z" or list[25]=="z":
                                                                                                                                                                                                                        if n>26:
                                                                                                                                                                                                                            print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                                                            quit()
                                                                                                                                                                                                                        else:
                                                                                                                                                                                                                            print("\nSuccess..")
                                                                                                                                                                                                                            e=time.time()
                                                                                                                                                                                                                            t=e-s
                                                                                                                                                                                                                            print("Time is : ",e-s)
                                                                                                                                                                                                                            print("WPM : ",(26/5)/((e-s)/60))
                                                                                                                                                                                                                            print()
                                                                                                                                                                                                                            ld(name,t)
                                                                                                                                                                                                                            exit()
                                                                                                                                                                                                                    else:
                                                                                                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                                                        quit()
                                                                                                                                                                                                            else:
                                                                                                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                                e=time.time()
                                                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                                                quit()
                                                                                                                                                                                                    else:
                                                                                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                                        quit()
                                                                                                                                                                                            else:
                                                                                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                                                                                e=time.time()
                                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                                quit()
                                                                                                                                                                                    else:
                                                                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                                                                        e=time.time()
                                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                                        quit()
                                                                                                                                                                            else:
                                                                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                                                                e=time.time()
                                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                                quit()
                                                                                                                                                                    else:
                                                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                                                        e=time.time()
                                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                                        quit()
                                                                                                                                                            else:
                                                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                                                e=time.time()
                                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                                quit()
                                                                                                                                                    else:
                                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                                        e=time.time()
                                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                                        quit()
                                                                                                                                            else:
                                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                                e=time.time()
                                                                                                                                                print("Time is : ",e-s)
                                                                                                                                                quit()
                                                                                                                                    else:
                                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                                        e=time.time()
                                                                                                                                        print("Time is : ",e-s)
                                                                                                                                        quit()
                                                                                                                            else:
                                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                                e=time.time()
                                                                                                                                print("Time is : ",e-s)
                                                                                                                                quit()
                                                                                                                    else:
                                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                                        e=time.time()
                                                                                                                        print("Time is : ",e-s)
                                                                                                                        quit()
                                                                                                            else:
                                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                                e=time.time()
                                                                                                                print("Time is : ",e-s)
                                                                                                                quit()
                                                                                                    else:
                                                                                                        print("\nWRONG CAHARCHTER")
                                                                                                        e=time.time()
                                                                                                        print("Time is : ",e-s)
                                                                                                        quit()
                                                                                            else:
                                                                                                print("\nWRONG CAHARCHTER")
                                                                                                e=time.time()
                                                                                                print("Time is : ",e-s)
                                                                                                quit()
                                                                                    else:
                                                                                        print("\nWRONG CAHARCHTER")
                                                                                        e=time.time()
                                                                                        print("Time is : ",e-s)
                                                                                        quit()
                                                                            else:
                                                                                print("\nWRONG CAHARCHTER")
                                                                                e=time.time()
                                                                                print("Time is : ",e-s)
                                                                                quit()
                                                                    else:
                                                                        print("\nWRONG CAHARCHTER")
                                                                        e=time.time()
                                                                        print("Time is : ",e-s)
                                                                        quit()
                                                            else:
                                                                print("\nWRONG CAHARCHTER")
                                                                e=time.time()
                                                                print("Time is : ",e-s)
                                                                quit()
                                                    else:
                                                        print("\nWRONG CAHARCHTER")
                                                        e=time.time()
                                                        print("Time is : ",e-s)
                                                        quit()
                                            else:
                                                print("\nWRONG CAHARCHTER")
                                                e=time.time()
                                                print("Time is : ",e-s)
                                                quit()
                                    else:
                                        print("\nWRONG CAHARCHTER")
                                        e=time.time()
                                        print("Time is : ",e-s)
                                        quit()
                            else:
                                print("\nWRONG CAHARCHTER")
                                e=time.time()
                                print("Time is : ",e-s)
                                quit()
                    else:
                        print("\nWRONG CAHARCHTER")
                        e=time.time()
                        print("Time is : ",e-s)
                        quit()
            else:
                print("\nWRONG CAHARCHTER")
                e=time.time()
                print("Time is : ",e-s)
                quit()
    else:
        
        print("\nINVALID INPUT")
        
def ld(n1,t1):

    name=n1
    time=t1
    dict={}
    sdict={}
    namelist=[]
    timelist=[]
    
    ob=openpyxl.load_workbook("TypeBoard.xlsx")
    ba=ob.active
    
    for i in range(1,100):
        if ba.cell(row=i,column=1).value is None and ba.cell(row=i,column=2).value is None:
            ba.cell(row=i,column=1).value=name
            ba.cell(row=i,column=2).value=time
            break
    
    ob.save(filename="TypeBoard.xlsx")
    
    data=pd.read_excel("TypeBoard.xlsx")
    namelist=data["Name"].tolist()
    timelist=data["Time"].tolist()

    for i in range(len(namelist)):
        dict[namelist[i]]=timelist[i]
    
    sortv=sorted(dict.values())
    
    for i in sortv:
        for j in dict.keys():
            if dict[j]==i:
                sdict[j]=dict[j]
                break
                
    print("\nTOP 5\n")
    
    i=0
    for key in sdict:
        if i!=5:
            print(key,":",sdict[key])
            i=i+1    
